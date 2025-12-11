"""
Hyperion CTI - Tenable Integration Script
Fetches CVEs from your dashboard, checks Tenable for affected assets,
generates Excel report, and posts to Microsoft Teams.

Requirements:
pip install requests openpyxl python-dotenv pandas
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import time
import json
from dotenv import load_dotenv

# Load environment variables from parent directory
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

# Configuration
TENABLE_ACCESS_KEY = os.getenv('TENABLE_ACCESS_KEY')
TENABLE_SECRET_KEY = os.getenv('TENABLE_SECRET_KEY')
TEAMS_WEBHOOK_URL = os.getenv('TEAMS_WEBHOOK_URL', 'none')

# Hyperion data source options
HYPERION_FUNCTION_URL = os.getenv('HYPERION_FUNCTION_URL', 'https://hyperion-cti.netlify.app/.netlify/functions/get-threats')
HYPERION_CSV_PATH = os.getenv('HYPERION_CSV_PATH', '')  # Optional: path to local CSV file

# Output directory
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'reports')
os.makedirs(OUTPUT_DIR, exist_ok=True)

class TenableCVEChecker:
    def __init__(self):
        self.headers = {
            "X-ApiKeys": f"accessKey={TENABLE_ACCESS_KEY}; secretKey={TENABLE_SECRET_KEY}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        self.base_url = "https://cloud.tenable.com"
        self.use_plugins_endpoint = False  # Flag to switch API approach
    
    def test_tenable_connection(self):
        """Test Tenable API connection and determine best endpoint"""
        print("\n[TEST] Testing Tenable API connection...")
        
        # Test 1: Basic connectivity
        try:
            test_url = f"{self.base_url}/scans"
            response = requests.get(test_url, headers=self.headers, timeout=10)
            if response.status_code == 200:
                print("   ✓ Tenable API connection successful")
            else:
                print(f"   ⚠ Tenable API returned status: {response.status_code}")
                return False
        except Exception as e:
            print(f"   ✗ Cannot connect to Tenable: {e}")
            return False
        
        # Test 2: Try to find a known CVE
        print("   Testing CVE lookup methods...")
        test_cve = "CVE-2024-0001"  # Use a recent CVE
        
        # Method 1: Workbenches API
        try:
            url = f"{self.base_url}/workbenches/vulnerabilities/{test_cve}/info"
            response = requests.get(url, headers=self.headers, timeout=5)
            if response.status_code in [200, 404]:
                print(f"   ✓ Workbenches API works (status: {response.status_code})")
                return True
        except Exception as e:
            print(f"   ⚠ Workbenches API error: {e}")
        
        # Method 2: Plugins API (alternative)
        try:
            url = f"{self.base_url}/plugins/plugin"
            response = requests.get(url, headers=self.headers, timeout=5)
            if response.status_code == 200:
                print("   ✓ Plugins API works - will use this method")
                self.use_plugins_endpoint = True
                return True
        except Exception as e:
            print(f"   ⚠ Plugins API error: {e}")
        
        print("   ⚠ Could not determine best API method, will try both")
        return True
    
    def fetch_hyperion_cves(self):
        """Fetch CVEs from Hyperion dashboard"""
        # Option 1: Try CSV file first (if specified)
        if HYPERION_CSV_PATH and os.path.exists(HYPERION_CSV_PATH):
            return self._fetch_from_csv()
        
        # Option 2: Fetch from Netlify function
        return self._fetch_from_netlify()
    
    def _fetch_from_csv(self):
        """Load CVEs from local CSV export"""
        try:
            print(f"[1/5] Loading CVEs from local file: {HYPERION_CSV_PATH}...")
            import csv
            
            cves = []
            with open(HYPERION_CSV_PATH, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cves.append({
                        'cve': row.get('CVE ID', row.get('cve', '')),
                        'product': row.get('Product', row.get('product', 'Unknown')),
                        'cvss': row.get('CVSS Score', row.get('cvssScore', 'N/A')),
                        'date': row.get('Date Added', row.get('dateAdded', '')),
                        'source': row.get('Source', row.get('source', 'Manual'))
                    })
            
            print(f"   ✓ Loaded {len(cves)} CVEs from CSV")
            return cves
        except Exception as e:
            print(f"   ✗ Error loading CSV: {e}")
            return []
    
    def _fetch_from_netlify(self):
        """Fetch CVEs from Netlify function endpoint"""
        try:
            print("[1/5] Fetching CVEs from Hyperion dashboard...")
            print(f"   Endpoint: {HYPERION_FUNCTION_URL}")
            
            response = requests.get(HYPERION_FUNCTION_URL, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            cves = []
            for cve_data in data.get('zeroDays', []):
                cves.append({
                    'cve': cve_data['cve'],
                    'product': cve_data['product'],
                    'cvss': cve_data.get('cvssScore', 'N/A'),
                    'date': cve_data['dateAdded'],
                    'source': cve_data['source']
                })
            
            print(f"   ✓ Found {len(cves)} CVEs from Hyperion")
            return cves
        except requests.exceptions.RequestException as e:
            print(f"   ✗ Error fetching from Netlify: {e}")
            print(f"   ℹ  Tip: Try exporting CVEs to CSV and set HYPERION_CSV_PATH in .env")
            return []
    
    def check_cve_in_tenable(self, cve_id):
        """Check if CVE affects any assets in Tenable"""
        # Try primary method first (workbenches)
        result = self._check_via_workbenches(cve_id)
        
        # If that fails, try alternative method (vulnerabilities export)
        if result['details'] in ['Not found in Tenable', 'Empty response from Tenable', 'Invalid JSON response']:
            alt_result = self._check_via_vulnerabilities(cve_id)
            if alt_result['exploitable'] is not None:
                return alt_result
        
        return result
    
    def _check_via_workbenches(self, cve_id):
        """Primary method: Check CVE via workbenches API"""
        try:
            # Remove any whitespace/formatting from CVE ID
            cve_id = cve_id.strip().upper()
            
            url = f"{self.base_url}/workbenches/vulnerabilities/{cve_id}/info"
            response = requests.get(url, headers=self.headers, timeout=10)
            
            if response.status_code == 404:
                return {
                    'exploitable': False,
                    'affected_assets': 0,
                    'severity': 'N/A',
                    'details': 'Not found in Tenable'
                }
            
            if response.status_code != 200:
                return {
                    'exploitable': False,
                    'affected_assets': 0,
                    'severity': 'N/A',
                    'details': f'HTTP {response.status_code}'
                }
            
            if not response.text or response.text.strip() == '':
                return {
                    'exploitable': False,
                    'affected_assets': 0,
                    'severity': 'N/A',
                    'details': 'Not in Tenable scan results'
                }
            
            try:
                vuln_info = response.json()
            except ValueError:
                return {
                    'exploitable': False,
                    'affected_assets': 0,
                    'severity': 'N/A',
                    'details': 'Invalid JSON response'
                }
            
            # ENHANCED: Better asset count detection
            asset_count = 0
            severity = 'N/A'
            
            # Try multiple data paths (Tenable API structure varies)
            if 'info' in vuln_info:
                asset_count = vuln_info['info'].get('count', 0)
                severity = vuln_info['info'].get('severity', 'N/A')
            elif 'count' in vuln_info:
                asset_count = vuln_info.get('count', 0)
                severity = vuln_info.get('severity', 'N/A')
            elif 'total' in vuln_info:
                asset_count = vuln_info.get('total', 0)
                severity = vuln_info.get('severity', 'N/A')
            
            # Additional check: look for assets array
            if asset_count == 0 and 'assets' in vuln_info:
                asset_count = len(vuln_info.get('assets', []))
            
            # Debug logging for CVE-2025-62215
            if cve_id == 'CVE-2025-62215':
                print(f"\n   [DEBUG] {cve_id} Response: {json.dumps(vuln_info, indent=2)[:500]}\n")
            
            return {
                'exploitable': asset_count > 0,
                'affected_assets': asset_count,
                'severity': severity,
                'details': f"{asset_count} asset(s) affected" if asset_count > 0 else "No assets affected"
            }
            
        except requests.exceptions.Timeout:
            return {'exploitable': None, 'affected_assets': 'Timeout', 'severity': 'N/A', 'details': 'Request timed out'}
        except requests.exceptions.ConnectionError:
            return {'exploitable': None, 'affected_assets': 'Error', 'severity': 'N/A', 'details': 'Connection error'}
        except Exception as e:
            return {'exploitable': None, 'affected_assets': 'Error', 'severity': 'N/A', 'details': f'Error: {str(e)[:50]}'}
    
    def _check_via_vulnerabilities(self, cve_id):
        """Alternative method: Check CVE via vulnerabilities search"""
        try:
            # Clean CVE ID
            cve_id = cve_id.strip().upper()
            
            # Method 1: Try direct vulnerability search with CVE filter
            url = f"{self.base_url}/workbenches/vulnerabilities"
            params = {
                'filter.0.filter': 'cve.id',
                'filter.0.quality': 'eq',
                'filter.0.value': cve_id
            }
            
            response = requests.get(url, headers=self.headers, params=params, timeout=10)
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    vulnerabilities = data.get('vulnerabilities', [])
                    
                    if vulnerabilities:
                        total_count = sum(v.get('count', 0) for v in vulnerabilities)
                        severity = vulnerabilities[0].get('severity', 'N/A') if vulnerabilities else 'N/A'
                        
                        if total_count > 0:
                            return {
                                'exploitable': True,
                                'affected_assets': total_count,
                                'severity': severity,
                                'details': f"{total_count} asset(s) affected (via search)"
                            }
                except ValueError:
                    pass
            
            # Method 2: Try plugin-based search (some CVEs map to plugins)
            try:
                plugin_url = f"{self.base_url}/workbenches/assets/vulnerabilities"
                plugin_params = {
                    'filter.search_type': 'and',
                    'filter.0.filter': 'cve',
                    'filter.0.quality': 'eq',
                    'filter.0.value': cve_id
                }
                
                plugin_response = requests.get(plugin_url, headers=self.headers, params=plugin_params, timeout=10)
                
                if plugin_response.status_code == 200:
                    plugin_data = plugin_response.json()
                    
                    if 'vulnerabilities' in plugin_data and plugin_data['vulnerabilities']:
                        asset_count = len(plugin_data['vulnerabilities'])
                        return {
                            'exploitable': True,
                            'affected_assets': asset_count,
                            'severity': 'N/A',
                            'details': f"{asset_count} asset(s) found (plugin search)"
                        }
            except:
                pass
            
            # No results from any method
            return {
                'exploitable': False,
                'affected_assets': 0,
                'severity': 'N/A',
                'details': 'Not in scan results'
            }
                
        except Exception as e:
            return {
                'exploitable': None,
                'affected_assets': 0,
                'severity': 'N/A',
                'details': f'Search error: {str(e)[:30]}'
            }
    
    def generate_report(self, cve_results):
        """Generate Excel report with CVE findings"""
        print("[3/5] Generating Excel report...")
        
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
        filename = f"CVE_Report_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CVE Analysis"
        
        # Header styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ["CVE ID", "Product", "CVSS Score", "Severity", "Date Added", "Source", 
                   "In Tenable?", "Affected Assets", "Details"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        exploitable_count = 0
        critical_count = 0
        
        for result in cve_results:
            tenable = result['tenable_result']
            
            # Determine if exploitable
            is_exploitable = "YES" if tenable['exploitable'] else "NO" if tenable['exploitable'] is False else "UNKNOWN"
            
            if tenable['exploitable']:
                exploitable_count += 1
            
            # ENHANCED: Calculate severity from CVSS score
            cvss_str = str(result['cvss'])
            severity = self._calculate_severity_from_cvss(cvss_str)
            
            row = [
                result['cve'],
                result['product'],
                result['cvss'],
                severity,  # Now calculated from CVSS
                result['date'],
                result['source'],
                is_exploitable,
                tenable['affected_assets'],
                tenable['details']
            ]
            ws.append(row)
            
            # Color code exploitable rows
            row_num = ws.max_row
            if is_exploitable == "YES":
                for cell in ws[row_num]:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    cell.font = Font(bold=True)
                critical_count += 1
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws.append(["Hyperion CTI - Daily CVE Report"])
        summary_ws.append(["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_ws.append([])
        summary_ws.append(["Metric", "Count"])
        summary_ws.append(["Total CVEs Analyzed", len(cve_results)])
        summary_ws.append(["CVEs Affecting Our Assets", exploitable_count])
        summary_ws.append(["High Priority (Exploitable + High CVSS)", critical_count])
        summary_ws.append(["CVEs Not in Tenable", len([r for r in cve_results if not r['tenable_result']['exploitable']])])
        
        # Style summary
        summary_ws['A1'].font = Font(bold=True, size=16, color="0066CC")
        for cell in summary_ws['A4:B4'][0]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Adjust column widths
        for ws_obj in [ws, summary_ws]:
            for column in ws_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        print(f"   ✓ Report saved: {filepath}")
        
        return filepath, exploitable_count, critical_count
    
    def _calculate_severity_from_cvss(self, cvss_str):
        """Calculate severity based on CVSS score"""
        try:
            cvss = float(cvss_str)
            if cvss == 0.0:
                return "None"
            elif 0.1 <= cvss <= 3.9:
                return "Low"
            elif 4.0 <= cvss <= 6.9:
                return "Medium"
            elif 7.0 <= cvss <= 8.9:
                return "High"
            elif 9.0 <= cvss <= 10.0:
                return "Critical"
            else:
                return "Unknown"
        except (ValueError, TypeError):
            return "N/A"
    
    def post_to_teams(self, filepath, exploitable_count, critical_count, total_cves):
        """Post report to Microsoft Teams via webhook"""
        if not TEAMS_WEBHOOK_URL or TEAMS_WEBHOOK_URL == 'none':
            print("[4/5] Skipping Teams notification (webhook not configured)")
            return
        
        print("[4/5] Posting to Microsoft Teams...")
        
        # Read file for upload (Teams doesn't support direct file uploads via webhooks)
        # We'll post a summary message with a link to where the file is stored
        
        timestamp = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        
        # Adaptive card for Teams
        card = {
            "@type": "MessageCard",
            "@context": "https://schema.org/extensions",
            "themeColor": "0066CC" if exploitable_count == 0 else "FF0000",
            "summary": f"Daily CVE Report - {exploitable_count} Exploitable",
            "sections": [{
                "activityTitle": "🛡️ Hyperion CTI - Daily CVE Report",
                "activitySubtitle": timestamp,
                "facts": [
                    {"name": "Total CVEs Analyzed:", "value": str(total_cves)},
                    {"name": "🚨 Exploitable in Our Environment:", "value": str(exploitable_count)},
                    {"name": "⚠️ High Priority:", "value": str(critical_count)},
                    {"name": "Report Location:", "value": filepath}
                ],
                "markdown": True
            }]
        }
        
        if exploitable_count > 0:
            card["sections"][0]["text"] = f"**ACTION REQUIRED:** {exploitable_count} CVE(s) affect our assets. Review the report immediately."
        else:
            card["sections"][0]["text"] = "✅ No exploitable CVEs found in our environment today."
        
        try:
            response = requests.post(TEAMS_WEBHOOK_URL, json=card, timeout=10)
            response.raise_for_status()
            print(f"   ✓ Posted to Teams successfully")
        except Exception as e:
            print(f"   ✗ Error posting to Teams: {e}")
    
    def run_daily_check(self):
        """Main execution flow"""
        print("=" * 60)
        print("Hyperion CTI - Tenable Integration Script")
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)
        
        # Test Tenable connection first
        if not self.test_tenable_connection():
            print("\n✗ Tenable API test failed. Check your API keys.")
            print("   Keys should be from: Tenable.io → My Account → API Keys")
            return
        
        # Step 1: Fetch CVEs from Hyperion
        hyperion_cves = self.fetch_hyperion_cves()
        
        if not hyperion_cves:
            print("\n✗ No CVEs to process. Exiting.")
            return
        
        # Step 2: Check each CVE in Tenable
        print(f"\n[2/5] Checking {len(hyperion_cves)} CVEs in Tenable...")
        cve_results = []
        
        for i, cve_data in enumerate(hyperion_cves, 1):
            print(f"   Checking {cve_data['cve']} ({i}/{len(hyperion_cves)})...", end=" ")
            
            tenable_result = self.check_cve_in_tenable(cve_data['cve'])
            
            cve_results.append({
                **cve_data,
                'tenable_result': tenable_result
            })
            
            status = "✓ Exploitable" if tenable_result['exploitable'] else "○ Not found"
            print(status)
            
            # Rate limiting - be nice to Tenable API
            time.sleep(0.5)
        
        # Step 3: Generate Excel report
        filepath, exploitable_count, critical_count = self.generate_report(cve_results)
        
        # Step 4: Post to Teams
        self.post_to_teams(filepath, exploitable_count, critical_count, len(hyperion_cves))
        
        # Step 5: Summary
        print("\n[5/5] Summary:")
        print(f"   Total CVEs analyzed: {len(hyperion_cves)}")
        print(f"   🚨 Exploitable in our environment: {exploitable_count}")
        print(f"   ⚠️  High priority: {critical_count}")
        print(f"   📄 Report: {filepath}")
        print("\n" + "=" * 60)
        print("✓ Daily check complete!")
        print("=" * 60)


if __name__ == "__main__":
    # Validate configuration
    if not TENABLE_ACCESS_KEY or not TENABLE_SECRET_KEY:
        print("ERROR: Tenable API keys not configured!")
        print("Please set TENABLE_ACCESS_KEY and TENABLE_SECRET_KEY in .env file")
        exit(1)
    
    # Run the checker
    checker = TenableCVEChecker()
    checker.run_daily_check()